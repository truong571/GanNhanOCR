"""Xuat 2,075 mau matched=False ra file Excel co thumbnail anh crop de duyet tay.

    python evaluation/export_unmatched_review.py
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XImage
from openpyxl.styles import Font, PatternFill

REPO = Path(__file__).resolve().parent.parent
LABELS = REPO / "dataset" / "all" / "labels.csv"
CROPS_ROOT = REPO / "dataset"  # crops dung path relative `crops/page_X/...`
OUT = REPO / "evaluation" / "reports" / "unmatched_review.xlsx"


def main() -> None:
    df = pd.read_csv(LABELS)
    bad = df[~df["matched"]].copy()
    print(f"Unmatched: {len(bad)}")
    if not len(bad):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "unmatched"
    headers = ["#", "thumb", "source", "page", "nom_char", "unicode",
               "syllable", "tier", "crop_file", "manual_label"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(bold=True)
        ws.cell(1, c).fill = PatternFill("solid", fgColor="FFE4B5")
    ws.freeze_panes = "C2"
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["I"].width = 50
    ws.column_dimensions["J"].width = 16

    for i, (_, r) in enumerate(bad.iterrows(), start=2):
        ws.cell(i, 1, i - 1)
        ws.cell(i, 3, r["source"])
        ws.cell(i, 4, r["page"])
        ws.cell(i, 5, r["nom_char"])
        ws.cell(i, 5).font = Font(size=20, name="Noto Serif CJK SC")
        ws.cell(i, 6, r["unicode"])
        ws.cell(i, 7, r["syllable"])
        ws.cell(i, 8, r["tier"])
        ws.cell(i, 9, r["crop_file"])

        path = CROPS_ROOT / r["source"] / r["crop_file"]
        if path.exists():
            try:
                img = XImage(str(path))
                img.width, img.height = 64, 64
                ws.add_image(img, f"B{i}")
                ws.row_dimensions[i].height = 52
            except Exception:
                pass

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
