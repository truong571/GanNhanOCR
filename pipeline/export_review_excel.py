"""Xuất MẪU NGẪU NHIÊN sang Excel để soát tay: cột dữ liệu + ảnh crop nhúng thật.

Không xuất toàn bộ labels.csv (hàng chục nghìn dòng nhúng ảnh sẽ ra file quá
lớn, dễ treo Excel) — lấy mẫu ngẫu nhiên có seed cố định để tái lập được.

Usage:
    python3 pipeline/export_review_excel.py \
        --labels dataset/labels.csv --src-root dataset \
        --out dataset_out/review_sample_n846.xlsx --n 846 --seed 42
"""
from __future__ import annotations

import argparse
import csv
import random
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

COLUMNS = ["image", "book", "page", "ocr_char", "syllable",
           "label", "unicode", "label_level", "tier"]
THUMB_PX = 80


def build_excel(labels_path: Path, src_root: Path, out_path: Path, n: int, seed: int) -> int:
    if not labels_path.exists():
        print(f"[review-xlsx] không thấy {labels_path}", file=sys.stderr)
        return 1

    with open(labels_path, encoding="utf-8", newline="") as f:
        all_rows = [r for r in csv.DictReader(f) if r.get("image")]
    # chỉ giữ dòng CÓ NHÃN ký tự (label_level=char, GOLD/SILVER) — SYLLABLE
    # luôn label/unicode rỗng nên không có gì để soát nhãn.
    rows = [r for r in all_rows if r.get("label_level") == "char" and r.get("label")]
    if not rows:
        print(f"[review-xlsx] {labels_path} không có dòng char-level nào (label_level=char).",
              file=sys.stderr)
        return 1

    rng = random.Random(seed)
    n_take = min(n, len(rows))
    sample = rng.sample(rows, n_take)
    # sort lại cho dễ đối chiếu tay (random.sample không giữ thứ tự gốc)
    sample.sort(key=lambda r: (r["book"], r["page"], r["image"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "review_sample"
    header = COLUMNS + ["crop_image"]
    ws.append(header)
    for c in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    img_col_letter = get_column_letter(len(header))
    ws.column_dimensions[img_col_letter].width = THUMB_PX / 7.0

    n_missing = 0
    for i, r in enumerate(sample, start=2):
        ws.append([r.get(c, "") for c in COLUMNS] + [""])
        ws.row_dimensions[i].height = THUMB_PX * 0.75

        img_path = src_root / r["image"]
        if not img_path.exists():
            n_missing += 1
            continue
        pil_img = PILImage.open(img_path).convert("RGB")
        pil_img.thumbnail((THUMB_PX, THUMB_PX))
        buf = BytesIO()
        pil_img.save(buf, format="PNG")
        buf.seek(0)
        ws.add_image(XLImage(buf), f"{img_col_letter}{i}")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[review-xlsx] {out_path} — {len(sample)} dòng (yêu cầu n={n}, seed={seed}), "
          f"{n_missing} ảnh thiếu trên đĩa")
    return 0


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--labels", default="dataset/labels.csv")
    ap.add_argument("--src-root", default="dataset")
    ap.add_argument("--out", default="dataset_out/review_sample.xlsx")
    ap.add_argument("--n", type=int, default=846)
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()
    sys.exit(build_excel(Path(args.labels), Path(args.src_root), Path(args.out),
                          args.n, args.seed))


if __name__ == "__main__":
    main()
