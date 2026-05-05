"""Debug 1 column: xuat Excel chi tiet tung buoc pipeline.

Usage:
  python tools/debug_column.py CacThanhTruyen2 page_0012 2
  python tools/debug_column.py CacThanhTruyen2 page_0012 2 --out debug.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage

from pipeline.step0_setup import load_config


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("book", type=str)
    parser.add_argument("page", type=str)
    parser.add_argument("col", type=int)
    parser.add_argument("--out", type=str, default=None)
    parser.add_argument("--config", type=str, default="config/pipeline.yaml")
    args = parser.parse_args()

    config = load_config(args.config)
    data_dir = Path(config["paths"]["data_dir"]) / args.book
    col_num = args.col
    page = args.page
    out_path = args.out or f"debug_{args.book}_{page}_col{col_num:02d}.xlsx"

    # ── Load all data ──

    # 1. OCR API
    ocr_cache_path = data_dir / "detected" / f"{page}_ocr_cache.json"
    ocr_chars = []
    if ocr_cache_path.exists():
        cache = json.load(open(ocr_cache_path))
        cols = cache.get("columns", [])
        if col_num - 1 < len(cols):
            ocr_chars = cols[col_num - 1]

    # 2. Transcription QN
    trans_path = data_dir / "transcriptions" / f"{page}.txt"
    syllables = []
    if trans_path.exists():
        lines = trans_path.read_text(encoding="utf-8").strip().split("\n")
        if col_num - 1 < len(lines):
            syllables = lines[col_num - 1].split()

    # 3. Detection (after alignment + crop)
    det_path = data_dir / "detected" / f"{page}_detection.json"
    det_chars = []
    if det_path.exists():
        det = json.load(open(det_path))
        for c in det["columns"]:
            if c["column"] == col_num:
                det_chars = c["chars"]
                break

    # 4. Alignment
    aligned_path = data_dir / "aligned" / f"{page}_aligned.json"
    aligned_pairs = []
    if aligned_path.exists():
        all_aligned = json.load(open(aligned_path))
        aligned_pairs = [a for a in all_aligned if a.get("column") == col_num]

    # 5. Labels
    labeled_path = data_dir / "labeled" / "dataset.json"
    labels = []
    if labeled_path.exists():
        all_labels = json.load(open(labeled_path))
        labels = [l for l in all_labels
                  if l.get("column") == col_num and l["page"] == page]

    # ── Build Excel ──
    wb = Workbook()

    # ──────────────────────────────────────
    # Sheet 1: Tong quan pipeline
    # ──────────────────────────────────────
    ws = wb.active
    ws.title = "Tong quan"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    nom_font = Font(size=16, bold=True)
    match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    unmatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gap_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "#", "Crop", "OCR char", "QN syllable", "Align type",
        "Nom label", "Unicode", "Matched", "Tier", "Candidates",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row_idx = 2
    img_size = 50
    for i, pair in enumerate(aligned_pairs):
        atype = pair["type"]
        char_info = pair.get("char", {}) or {}
        syllable = pair.get("syllable", "")

        # Find corresponding label
        label = labels[i] if i < len(labels) else {}

        # #
        ws.cell(row=row_idx, column=1, value=i + 1).border = border

        # Crop image
        crop_file = char_info.get("crop_file", "")
        if crop_file:
            crop_path = data_dir / "detected" / crop_file
            if crop_path.exists():
                try:
                    img = XlImage(str(crop_path))
                    img.width = img_size
                    img.height = img_size
                    ws.add_image(img, f"B{row_idx}")
                except Exception:
                    pass
        ws.cell(row=row_idx, column=2).border = border

        # OCR char
        ocr = char_info.get("ocr_char", "")
        cell = ws.cell(row=row_idx, column=3, value=ocr or "")
        cell.font = Font(size=14)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

        # QN syllable
        ws.cell(row=row_idx, column=4, value=syllable or "").border = border

        # Align type
        cell = ws.cell(row=row_idx, column=5, value=atype)
        cell.border = border
        if atype == "match":
            cell.fill = match_fill
        elif atype == "insertion":
            cell.fill = gap_fill
        elif atype == "deletion":
            cell.fill = gap_fill

        # Nom label
        nom = label.get("nom_char", "")
        cell = ws.cell(row=row_idx, column=6, value=nom or "")
        cell.font = nom_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

        # Unicode
        uni = label.get("unicode", "")
        ws.cell(row=row_idx, column=7, value=uni or "").border = border

        # Matched
        matched = label.get("matched", False)
        cell = ws.cell(row=row_idx, column=8, value=str(matched))
        cell.border = border
        if matched:
            cell.fill = match_fill
            cell.font = Font(color="008000", bold=True)
        else:
            cell.fill = unmatch_fill
            cell.font = Font(color="FF0000", bold=True)

        # Tier
        tier = label.get("tier", "")
        cell = ws.cell(row=row_idx, column=9, value=tier)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

        # Candidates
        cands = label.get("nom_candidates", [])
        ws.cell(row=row_idx, column=10, value=" ".join(cands[:10]) if cands else "").border = border

        ws.row_dimensions[row_idx].height = img_size * 0.75 + 5
        row_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 6
    ws.column_dimensions["J"].width = 30
    ws.freeze_panes = "A2"

    # ──────────────────────────────────────
    # Sheet 2: OCR API raw
    # ──────────────────────────────────────
    ws2 = wb.create_sheet("OCR API")
    ocr_headers = ["#", "Char", "y_center", "bbox"]
    for c, h in enumerate(ocr_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for i, oc in enumerate(ocr_chars):
        ws2.cell(row=i+2, column=1, value=i+1)
        ws2.cell(row=i+2, column=2, value=oc["char"]).font = Font(size=14)
        ws2.cell(row=i+2, column=3, value=round(oc["y_center"]))
        ws2.cell(row=i+2, column=4, value=str(oc["bbox"]))

    # ──────────────────────────────────────
    # Sheet 3: QN Transcription
    # ──────────────────────────────────────
    ws3 = wb.create_sheet("QN Syllables")
    ws3.cell(row=1, column=1, value="#").font = header_font
    ws3.cell(row=1, column=1).fill = header_fill
    ws3.cell(row=1, column=2, value="Syllable").font = header_font
    ws3.cell(row=1, column=2).fill = header_fill
    for i, s in enumerate(syllables):
        ws3.cell(row=i+2, column=1, value=i+1)
        ws3.cell(row=i+2, column=2, value=s)

    # ──────────────────────────────────────
    # Sheet 4: So sanh
    # ──────────────────────────────────────
    ws4 = wb.create_sheet("So sanh OCR vs QN")
    cmp_headers = ["#", "OCR char", "QN syllable", "Align", "Ghi chu"]
    for c, h in enumerate(cmp_headers, 1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill

    max_len = max(len(ocr_chars), len(syllables))
    for i in range(max_len):
        row = i + 2
        ws4.cell(row=row, column=1, value=i+1)
        if i < len(ocr_chars):
            ws4.cell(row=row, column=2, value=ocr_chars[i]["char"]).font = Font(size=14)
        if i < len(syllables):
            ws4.cell(row=row, column=3, value=syllables[i])

    # Fill alignment info
    for i, pair in enumerate(aligned_pairs):
        row = i + 2
        ws4.cell(row=row, column=4, value=pair["type"])
        if pair["type"] == "insertion":
            ws4.cell(row=row, column=5, value="QN thua, khong co anh")
            ws4.cell(row=row, column=4).fill = gap_fill
        elif pair["type"] == "deletion":
            ws4.cell(row=row, column=5, value="Anh thua, khong co QN")
            ws4.cell(row=row, column=4).fill = gap_fill

    # ──────────────────────────────────────
    # Sheet 5: Info
    # ──────────────────────────────────────
    ws5 = wb.create_sheet("Info")
    info = [
        ("Book", args.book),
        ("Page", page),
        ("Column", col_num),
        ("OCR chars", len(ocr_chars)),
        ("QN syllables", len(syllables)),
        ("Aligned pairs", len(aligned_pairs)),
        ("  Matches", sum(1 for a in aligned_pairs if a["type"] == "match")),
        ("  Insertions", sum(1 for a in aligned_pairs if a["type"] == "insertion")),
        ("  Deletions", sum(1 for a in aligned_pairs if a["type"] == "deletion")),
        ("Labels matched", sum(1 for l in labels if l.get("matched"))),
        ("Labels unmatched", sum(1 for l in labels if not l.get("matched") and l["type"] == "match")),
        ("Tier 1 (dict)", sum(1 for l in labels if l.get("tier") == 1)),
        ("Tier 2 (similar)", sum(1 for l in labels if l.get("tier") == 2)),
        ("Tier 3 (visual)", sum(1 for l in labels if l.get("tier") == 3)),
        ("Tier 0 (none)", sum(1 for l in labels if l.get("tier") == 0)),
    ]
    for i, (k, v) in enumerate(info, 1):
        ws5.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws5.cell(row=i, column=2, value=v)

    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"  Sheet 'Tong quan': pipeline result + crop images")
    print(f"  Sheet 'OCR API': raw OCR chars + bbox")
    print(f"  Sheet 'QN Syllables': QN transcription")
    print(f"  Sheet 'So sanh OCR vs QN': alignment")
    print(f"  Sheet 'Info': thong ke")


if __name__ == "__main__":
    main()
